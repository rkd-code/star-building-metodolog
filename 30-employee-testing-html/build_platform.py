#!/usr/bin/env python3
"""Собирает автономную HTML-платформу из Excel-банка."""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
BANK = ROOT.parent / "29-employee-testing-service" / "question_bank_pool_001.xlsx"
OUTPUT = ROOT / "index.html"


def load_bank(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=False, read_only=False)
    variants: dict[str, list[dict]] = defaultdict(list)
    for code, letter, text, correct, weight, order in wb["Варианты"].iter_rows(min_row=2, values_only=True):
        variants[code].append({
            "id": letter,
            "text": text,
            "correct": correct == "Да",
            "order": int(order),
        })
    roles: dict[str, list[str]] = defaultdict(list)
    for code, role, required, comment in wb["Назначение_по_должности"].iter_rows(min_row=2, values_only=True):
        roles[code].append(role)
    questions = []
    for row in wb["Вопросы"].iter_rows(min_row=2, values_only=True):
        code, topic, qtype, difficulty, wording, explanation, source, clause, status, version, author, reviewer, reviewed_at = row
        questions.append({
            "id": code,
            "topic": topic,
            "type": qtype,
            "difficulty": difficulty,
            "text": wording,
            "explanation": explanation,
            "source": source,
            "clause": clause,
            "status": status,
            "version": version,
            "roles": sorted(roles[code]),
            "options": sorted(variants[code], key=lambda option: option["order"]),
        })
    return questions


def score_answers(correctness: list[bool], threshold: int = 90) -> dict:
    total = len(correctness)
    correct = sum(bool(value) for value in correctness)
    percent = round((correct / total * 100) if total else 0, 1)
    return {"correct": correct, "total": total, "percent": percent, "passed": percent >= threshold}


def can_attempt(last_attempt_iso: str | None, current_iso: str) -> bool:
    if not last_attempt_iso:
        return True
    last = datetime.fromisoformat(last_attempt_iso)
    current = datetime.fromisoformat(current_iso)
    return (last.year, last.month) != (current.year, current.month)


HTML_TEMPLATE = r'''<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="light">
<title>Экзаменатор Star Building</title>
<style>
:root{--ink:#132329;--muted:#65757b;--paper:#f5f2ea;--card:#fffdfa;--line:#d8ddd9;--green:#1c6b50;--green2:#0f4d3a;--lime:#d9ef49;--red:#b24335;--amber:#d38b24;--shadow:0 18px 50px rgba(20,49,42,.12);--radius:22px}
*{box-sizing:border-box}body{margin:0;background:var(--paper);color:var(--ink);font:16px/1.5 Inter,Segoe UI,Arial,sans-serif}button,input,select{font:inherit}.hidden{display:none!important}
.shell{min-height:100vh;background:radial-gradient(circle at 82% 8%,rgba(217,239,73,.23),transparent 26%),linear-gradient(135deg,#f7f4ec 0%,#edf2ed 100%)}
.topbar{display:flex;align-items:center;justify-content:space-between;padding:18px clamp(18px,4vw,64px);border-bottom:1px solid rgba(19,35,41,.1);background:rgba(255,253,250,.76);backdrop-filter:blur(16px);position:sticky;top:0;z-index:20}
.brand{display:flex;align-items:center;gap:13px;font-weight:800;letter-spacing:-.02em}.mark{width:42px;height:42px;border-radius:13px;background:var(--green2);display:grid;place-items:center;color:var(--lime);box-shadow:0 7px 20px rgba(15,77,58,.22)}.brand small{display:block;color:var(--muted);font-weight:600;letter-spacing:.04em;text-transform:uppercase;font-size:10px}
.mode{display:flex;gap:6px;padding:5px;border:1px solid var(--line);border-radius:14px;background:#fff}.mode button{border:0;background:transparent;padding:9px 13px;border-radius:10px;cursor:pointer;color:var(--muted);font-weight:700}.mode button.active{background:var(--green2);color:white}
main{width:min(1180px,calc(100% - 32px));margin:0 auto;padding:42px 0 72px}.hero{display:grid;grid-template-columns:1.2fr .8fr;gap:26px;align-items:stretch}.panel{background:var(--card);border:1px solid rgba(19,35,41,.11);border-radius:var(--radius);box-shadow:var(--shadow)}.intro{padding:clamp(28px,5vw,58px)}
.eyebrow{display:inline-flex;gap:8px;align-items:center;color:var(--green);font-weight:800;text-transform:uppercase;letter-spacing:.12em;font-size:12px}.dot{width:8px;height:8px;border-radius:99px;background:var(--lime);box-shadow:0 0 0 5px rgba(217,239,73,.28)}h1{font-size:clamp(42px,6vw,76px);line-height:.96;letter-spacing:-.065em;margin:22px 0 20px;max-width:760px}.lead{font-size:19px;color:var(--muted);max-width:650px}.facts{display:flex;gap:10px;flex-wrap:wrap;margin-top:28px}.pill{border:1px solid var(--line);border-radius:999px;padding:9px 13px;background:#fff;font-weight:700;font-size:13px}
.login{padding:28px;display:flex;flex-direction:column}.login h2,.dashboard h2{margin:0 0 6px;font-size:25px;letter-spacing:-.03em}.note{color:var(--muted);font-size:13px}.field{display:grid;gap:7px;margin-top:17px}.field label{font-size:13px;font-weight:800}.field input,.field select{width:100%;border:1px solid var(--line);border-radius:12px;padding:13px 14px;background:white;color:var(--ink);outline:none}.field input:focus,.field select:focus{border-color:var(--green);box-shadow:0 0 0 3px rgba(28,107,80,.12)}
.btn{border:0;border-radius:13px;padding:13px 18px;font-weight:800;cursor:pointer;transition:.2s transform,.2s opacity,.2s background}.btn:hover{transform:translateY(-1px)}.btn.primary{background:var(--green2);color:#fff}.btn.primary:hover{background:var(--green)}.btn.secondary{background:#fff;border:1px solid var(--line);color:var(--ink)}.btn.danger{background:#fff0ed;color:var(--red);border:1px solid #efc9c2}.btn:disabled{opacity:.42;cursor:not-allowed;transform:none}.login .btn{margin-top:auto;margin-bottom:0}.alert{padding:13px 15px;border-radius:12px;margin-top:14px;font-size:13px;font-weight:650}.alert.warn{background:#fff4dd;color:#7a5419}.alert.ok{background:#eaf6ef;color:#175b43}.alert.error{background:#fff0ed;color:#8b3328}
.exam{display:grid;grid-template-columns:260px 1fr;gap:22px}.side{padding:22px;align-self:start;position:sticky;top:96px}.metric{padding:14px 0;border-bottom:1px solid var(--line)}.metric:last-child{border:0}.metric strong{display:block;font-size:25px;letter-spacing:-.04em}.progress{height:9px;background:#e8ece8;border-radius:99px;overflow:hidden;margin-top:12px}.progress i{display:block;height:100%;background:linear-gradient(90deg,var(--green),var(--lime));width:0;transition:.3s}.question{padding:clamp(24px,4vw,44px)}.qmeta{display:flex;justify-content:space-between;gap:14px;color:var(--muted);font-size:13px;font-weight:700}.question h2{font-size:clamp(24px,3vw,38px);line-height:1.18;letter-spacing:-.035em;margin:25px 0}.options{display:grid;gap:12px}.option{display:grid;grid-template-columns:34px 1fr;align-items:center;gap:12px;border:1px solid var(--line);border-radius:16px;padding:15px;background:white;cursor:pointer;transition:.15s}.option:hover{border-color:#9bb3a8;transform:translateX(2px)}.option.selected{border-color:var(--green);background:#eef7f1;box-shadow:0 0 0 2px rgba(28,107,80,.1)}.option b{width:32px;height:32px;border-radius:10px;background:#edf0ec;display:grid;place-items:center}.option.selected b{background:var(--green2);color:#fff}.nav{display:flex;justify-content:space-between;gap:12px;margin-top:28px}.qgrid{display:grid;grid-template-columns:repeat(6,1fr);gap:7px;margin-top:14px}.qgrid button{aspect-ratio:1;border:1px solid var(--line);border-radius:8px;background:white;font-size:11px;cursor:pointer}.qgrid button.done{background:var(--green2);color:white}.qgrid button.current{outline:3px solid var(--lime)}
.result{padding:clamp(28px,5vw,58px)}.score{display:flex;align-items:center;gap:25px;margin:24px 0}.score-ring{width:150px;height:150px;border-radius:50%;display:grid;place-items:center;background:conic-gradient(var(--green) var(--score),#e4e8e4 0);position:relative}.score-ring:after{content:"";position:absolute;inset:14px;border-radius:50%;background:var(--card)}.score-ring strong{position:relative;z-index:1;font-size:34px}.status{font-size:35px;font-weight:900;letter-spacing:-.04em}.topics{display:grid;gap:10px;margin-top:22px}.topic-row{display:grid;grid-template-columns:minmax(180px,1fr) 2fr 70px;gap:12px;align-items:center}.bar{height:10px;border-radius:99px;background:#e6e9e5;overflow:hidden}.bar i{display:block;height:100%;background:var(--green)}
.dashboard{padding:30px}.dash-head{display:flex;justify-content:space-between;gap:18px;align-items:flex-start}.actions{display:flex;gap:8px;flex-wrap:wrap}.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:24px 0}.stat{padding:18px;border:1px solid var(--line);border-radius:16px;background:#fff}.stat span{color:var(--muted);font-size:12px;font-weight:800;text-transform:uppercase}.stat strong{display:block;font-size:28px;margin-top:4px}.settings{display:grid;grid-template-columns:1fr 1fr auto;gap:12px;align-items:end;padding:18px;background:#f0f4ef;border-radius:16px;margin-bottom:20px}.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:16px}table{width:100%;border-collapse:collapse;min-width:780px}th,td{text-align:left;padding:13px 14px;border-bottom:1px solid var(--line);font-size:13px}th{background:#f3f5f1;color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.06em}tr:last-child td{border-bottom:0}.badge{display:inline-block;padding:5px 9px;border-radius:99px;font-size:11px;font-weight:800}.badge.pass{background:#e5f4eb;color:#126044}.badge.fail{background:#ffebe7;color:#97392d}.empty{text-align:center;padding:48px;color:var(--muted)}.footnote{margin-top:18px;padding:13px;border-radius:12px;background:#fff4dd;color:#74501b;font-size:12px}
@media(max-width:850px){.hero,.exam{grid-template-columns:1fr}.side{position:static}.cards{grid-template-columns:1fr 1fr}.settings{grid-template-columns:1fr}.topbar{align-items:flex-start}.brand div:last-child{display:none}}@media(max-width:520px){main{width:min(100% - 20px,1180px);padding-top:20px}.topbar{padding:12px}.mode button{padding:8px;font-size:12px}.cards{grid-template-columns:1fr}.question{padding:20px}.topic-row{grid-template-columns:1fr 60px}.topic-row .bar{grid-row:2;grid-column:1/3}.score{align-items:flex-start;flex-direction:column}.qgrid{grid-template-columns:repeat(8,1fr)}}
@media print{.topbar,.nav,.side,.actions,.settings,.footnote{display:none!important}.panel{box-shadow:none}.shell{background:white}main{width:100%;padding:0}}
</style>
</head>
<body><div class="shell">
<header class="topbar"><div class="brand"><div class="mark">SB</div><div>Экзаменатор<small>Star Building · отдел продаж</small></div></div><div class="mode"><button id="employeeMode" class="active">Сотрудник</button><button id="ropMode">Кабинет РОПа</button></div></header>
<main>
<section id="startView" class="hero">
<div class="panel intro"><span class="eyebrow"><i class="dot"></i>Проверка знаний</span><h1>Знать.<br>Применять.<br>Строить.</h1><p class="lead">Корпоративная проверка знаний технологий, продукта, регламентов и продаж Star Building.</p><div class="facts"><span class="pill">Банк: 102 вопроса</span><span class="pill">Порог: 90%</span><span class="pill">1 попытка в месяц</span><span class="pill">Результат — до запроса РОПа</span></div></div>
<div class="panel login"><h2>Начать тестирование</h2><p class="note">Укажите данные сотрудника. Прогресс сохраняется на этом устройстве.</p><div class="field"><label for="employeeId">Табельный номер или рабочая почта</label><input id="employeeId" autocomplete="username" placeholder="Например: sales-014"></div><div class="field"><label for="employeeName">Фамилия и имя</label><input id="employeeName" autocomplete="name" placeholder="Иван Петров"></div><div class="field"><label for="employeeRole">Должность</label><select id="employeeRole"><option>Менеджеры продаж</option><option>РОП / Руководитель отдела продаж</option></select></div><div id="startMessage"></div><button id="startButton" class="btn primary">Начать тест</button></div>
</section>
<section id="examView" class="exam hidden"><aside class="panel side"><div class="metric"><span class="note">Сотрудник</span><strong id="sideName">—</strong></div><div class="metric"><span class="note">Прогресс</span><strong id="progressText">0 / 0</strong><div class="progress"><i id="progressBar"></i></div></div><div id="timeMetric" class="metric hidden"><span class="note">Осталось времени</span><strong id="timeText">—</strong></div><div class="metric"><span class="note">До проходного результата</span><strong>90%</strong></div><div id="questionGrid" class="qgrid"></div></aside><article class="panel question"><div class="qmeta"><span id="qTopic"></span><span id="qDifficulty"></span></div><h2 id="qText"></h2><div id="options" class="options"></div><div class="nav"><button id="prevButton" class="btn secondary">Назад</button><button id="nextButton" class="btn primary">Далее</button></div></article></section>
<section id="resultView" class="panel result hidden"><span class="eyebrow"><i class="dot"></i>Результат зафиксирован</span><div class="score"><div id="scoreRing" class="score-ring"><strong id="scorePercent">0%</strong></div><div><div id="resultStatus" class="status"></div><p id="resultSummary" class="lead"></p></div></div><h2>Результат по темам</h2><div id="topicResults" class="topics"></div><div class="nav"><button class="btn secondary" onclick="window.print()">Распечатать</button><button class="btn primary" onclick="showStart()">На главную</button></div></section>
<section id="ropView" class="panel dashboard hidden"><div class="dash-head"><div><span class="eyebrow"><i class="dot"></i>Управленческий обзор</span><h2>Кабинет РОПа</h2><p class="note">Результаты, сохраненные в этом браузере.</p></div><div class="actions"><button id="exportCsv" class="btn secondary">Выгрузить результаты CSV</button><button id="exportJson" class="btn secondary">Резервная копия</button><label class="btn secondary" style="cursor:pointer">Загрузить копию<input id="importJson" type="file" accept="application/json" hidden></label></div></div><div class="cards"><div class="stat"><span>Сотрудников</span><strong id="statEmployees">0</strong></div><div class="stat"><span>Попыток</span><strong id="statAttempts">0</strong></div><div class="stat"><span>Прошли</span><strong id="statPassed">0%</strong></div><div class="stat"><span>Средний балл</span><strong id="statAverage">0%</strong></div></div><div class="settings"><div class="field" style="margin:0"><label>Вопросов в попытке</label><input id="settingCount" type="number" min="1" max="102" placeholder="102 — весь банк"></div><div class="field" style="margin:0"><label>Время, минут</label><input id="settingMinutes" type="number" min="1" placeholder="Без ограничения"></div><button id="saveSettings" class="btn primary">Сохранить</button></div><div class="table-wrap"><table><thead><tr><th>Сотрудник</th><th>Должность</th><th>Дата</th><th>Баллы</th><th>Статус</th><th>Действие</th></tr></thead><tbody id="resultsBody"></tbody></table><div id="emptyResults" class="empty">Результатов пока нет</div></div><div class="footnote">Автономная версия хранит данные только в браузере и не обеспечивает серверную авторизацию. Для корпоративного использования требуется подключение базы данных и учетных записей.</div></section>
</main></div>
<script id="question-data" type="application/json">__QUESTION_DATA__</script>
<script>
'use strict';
const BANK=JSON.parse(document.getElementById('question-data').textContent);const KEY='sb_exam_v1';const $=id=>document.getElementById(id);let state=loadState();let session=null;let timerHandle=null;
function defaults(){return{settings:{count:null,minutes:null,threshold:90},attempts:[],drafts:{},retest:{}}}function loadState(){try{return Object.assign(defaults(),JSON.parse(localStorage.getItem(KEY)||'{}'))}catch(e){return defaults()}}function saveState(){localStorage.setItem(KEY,JSON.stringify(state))}
function escapeText(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]))}function shuffle(a){a=[...a];for(let i=a.length-1;i>0;i--){const j=Math.floor(Math.random()*(i+1));[a[i],a[j]]=[a[j],a[i]]}return a}
function sameMonth(a,b){a=new Date(a);b=new Date(b);return a.getFullYear()===b.getFullYear()&&a.getMonth()===b.getMonth()}function lastAttempt(id){return state.attempts.filter(a=>a.employeeId===id).sort((a,b)=>b.finished.localeCompare(a.finished))[0]}
function eligibility(id){const last=lastAttempt(id),requested=state.retest[id];if(!last)return{ok:true};if(requested&&requested>last.finished){return sameMonth(last.finished,new Date())?{ok:false,msg:'РОП назначил повторное прохождение. Новая попытка будет доступна в следующем календарном месяце.'}:{ok:true}}return sameMonth(last.finished,new Date())?{ok:false,msg:'Попытка за текущий календарный месяц уже использована.'}:{ok:false,msg:'Результат действует. Новое прохождение откроется после запроса РОПа.'}}
function setView(id){if(id!=='examView'&&timerHandle){clearInterval(timerHandle);timerHandle=null}['startView','examView','resultView','ropView'].forEach(x=>$(x).classList.toggle('hidden',x!==id));$('employeeMode').classList.toggle('active',id!=='ropView');$('ropMode').classList.toggle('active',id==='ropView')}
function showStart(){setView('startView');session=null;checkEligibility()}function checkEligibility(){const id=$('employeeId').value.trim();const box=$('startMessage');box.innerHTML='';$('startButton').disabled=false;if(id){const e=eligibility(id);if(!e.ok){box.innerHTML='<div class="alert warn">'+escapeText(e.msg)+'</div>';$('startButton').disabled=true}}}
function startExam(){const employeeId=$('employeeId').value.trim(),name=$('employeeName').value.trim(),role=$('employeeRole').value;if(!employeeId||!name){$('startMessage').innerHTML='<div class="alert error">Заполните идентификатор и имя сотрудника.</div>';return}const e=eligibility(employeeId);if(!e.ok){checkEligibility();return}const count=Math.min(Number(state.settings.count)||BANK.length,BANK.length);const draft=state.drafts[employeeId];if(draft&&draft.questionIds?.length){session={...draft};}else{const picked=shuffle(BANK).slice(0,count),minutes=Number(state.settings.minutes)||null;session={employeeId,name,role,started:new Date().toISOString(),deadline:minutes?new Date(Date.now()+minutes*60000).toISOString():null,questionIds:picked.map(q=>q.id),optionOrder:Object.fromEntries(picked.map(q=>[q.id,shuffle(q.options.map(o=>o.id))])),answers:{},index:0};state.drafts[employeeId]=session;saveState()}$('sideName').textContent=session.name;setView('examView');renderQuestion();startTimer()}
function startTimer(){if(timerHandle){clearInterval(timerHandle);timerHandle=null}if(!session.deadline){$('timeMetric').classList.add('hidden');return}$('timeMetric').classList.remove('hidden');const tick=()=>{const seconds=Math.max(0,Math.ceil((new Date(session.deadline)-Date.now())/1000)),minutes=Math.floor(seconds/60),rest=String(seconds%60).padStart(2,'0');$('timeText').textContent=minutes+':'+rest;if(seconds<=0){clearInterval(timerHandle);timerHandle=null;finishExam(true)}};tick();if(session)timerHandle=setInterval(tick,1000)}
function questions(){return session.questionIds.map(id=>BANK.find(q=>q.id===id))}function renderQuestion(){const qs=questions(),q=qs[session.index];$('qTopic').textContent=q.topic;$('qDifficulty').textContent=q.difficulty;$('qText').textContent=q.text;const opts=session.optionOrder[q.id].map(id=>q.options.find(o=>o.id===id));$('options').innerHTML=opts.map(o=>'<div class="option '+(session.answers[q.id]===o.id?'selected':'')+'" data-id="'+o.id+'"><b>'+o.id+'</b><span>'+escapeText(o.text)+'</span></div>').join('');document.querySelectorAll('.option').forEach(el=>el.onclick=()=>answer(q.id,el.dataset.id));$('prevButton').disabled=session.index===0;$('nextButton').textContent=session.index===qs.length-1?'Завершить':'Далее';renderProgress()}
function answer(qid,oid){session.answers[qid]=oid;state.drafts[session.employeeId]=session;saveState();renderQuestion()}function renderProgress(){const total=session.questionIds.length,done=Object.keys(session.answers).length;$('progressText').textContent=done+' / '+total;$('progressBar').style.width=(done/total*100)+'%';$('questionGrid').innerHTML=session.questionIds.map((id,i)=>'<button class="'+(session.answers[id]?'done ':'')+(i===session.index?'current':'')+'" data-i="'+i+'">'+(i+1)+'</button>').join('');$('questionGrid').querySelectorAll('button').forEach(b=>b.onclick=()=>{session.index=Number(b.dataset.i);renderQuestion()})}
function move(step){if(step<0){session.index=Math.max(0,session.index-1);renderQuestion();return}if(session.index<session.questionIds.length-1){session.index++;renderQuestion()}else finishExam()}
function finishExam(timedOut=false){const qs=questions(),missing=qs.filter(q=>!session.answers[q.id]).length;if(missing&&!timedOut&&!confirm('Без ответа: '+missing+'. Завершить попытку?'))return;const details=qs.map(q=>{const selected=session.answers[q.id]||null,correct=q.options.find(o=>o.correct).id;return{id:q.id,topic:q.topic,selected,correct,isCorrect:selected===correct}});const correct=details.filter(x=>x.isCorrect).length,percent=Math.round(correct/details.length*1000)/10,passed=percent>=90;const record={id:'ATT-'+Date.now(),employeeId:session.employeeId,name:session.name,role:session.role,started:session.started,finished:new Date().toISOString(),percent,correct,total:details.length,passed,timedOut,details,valid:true};state.attempts.push(record);delete state.drafts[session.employeeId];delete state.retest[session.employeeId];saveState();renderResult(record)}
function renderResult(r){setView('resultView');$('scorePercent').textContent=r.percent+'%';$('scoreRing').style.setProperty('--score',r.percent+'%');$('resultStatus').textContent=r.passed?'Тест пройден':'Тест не пройден';$('resultStatus').style.color=r.passed?'var(--green)':'var(--red)';$('resultSummary').textContent=(r.timedOut?'Время истекло. ':'')+'Правильных ответов: '+r.correct+' из '+r.total+'. Проходной порог — 90%.';const grouped={};r.details.forEach(d=>{grouped[d.topic]??={all:0,ok:0};grouped[d.topic].all++;if(d.isCorrect)grouped[d.topic].ok++});$('topicResults').innerHTML=Object.entries(grouped).sort().map(([topic,v])=>{const p=Math.round(v.ok/v.all*100);return'<div class="topic-row"><span>'+escapeText(topic)+'</span><div class="bar"><i style="width:'+p+'%"></i></div><b>'+p+'%</b></div>'}).join('')}
function renderRop(){setView('ropView');const a=state.attempts,employees=new Set(a.map(x=>x.employeeId)).size,passed=a.length?a.filter(x=>x.passed).length/a.length*100:0,average=a.length?a.reduce((s,x)=>s+x.percent,0)/a.length:0;$('statEmployees').textContent=employees;$('statAttempts').textContent=a.length;$('statPassed').textContent=Math.round(passed)+'%';$('statAverage').textContent=Math.round(average*10)/10+'%';$('settingCount').value=state.settings.count||'';$('settingMinutes').value=state.settings.minutes||'';$('resultsBody').innerHTML=[...a].sort((x,y)=>y.finished.localeCompare(x.finished)).map(r=>'<tr><td><b>'+escapeText(r.name)+'</b><br><span class="note">'+escapeText(r.employeeId)+'</span></td><td>'+escapeText(r.role)+'</td><td>'+new Date(r.finished).toLocaleDateString('ru-RU')+'</td><td><b>'+r.percent+'%</b><br>'+r.correct+' / '+r.total+'</td><td><span class="badge '+(r.passed?'pass':'fail')+'">'+(r.passed?'Пройден':'Не пройден')+'</span></td><td><button class="btn secondary retest" data-id="'+escapeText(r.employeeId)+'">Запросить повтор</button></td></tr>').join('');$('emptyResults').classList.toggle('hidden',a.length>0);document.querySelectorAll('.retest').forEach(b=>b.onclick=()=>{state.retest[b.dataset.id]=new Date().toISOString();saveState();alert('Повторное прохождение назначено. Попытка откроется в следующем календарном месяце.')})}
function saveSettings(){const c=Number($('settingCount').value),m=Number($('settingMinutes').value);state.settings.count=c>=1&&c<=BANK.length?c:null;state.settings.minutes=m>=1?m:null;saveState();alert('Настройки сохранены. Пустое значение означает весь банк и отсутствие таймера.')}
function csv(){const rows=[['Идентификатор','Сотрудник','Должность','Дата','Правильно','Всего','Процент','Статус'],...state.attempts.map(a=>[a.employeeId,a.name,a.role,a.finished,a.correct,a.total,a.percent,a.passed?'Пройден':'Не пройден'])];download('results.csv','\ufeff'+rows.map(r=>r.map(x=>'"'+String(x).replace(/"/g,'""')+'"').join(';')).join('\n'),'text/csv')}
function download(name,text,type){const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([text],{type}));a.download=name;a.click();URL.revokeObjectURL(a.href)}function backup(){download('exam-backup.json',JSON.stringify(state,null,2),'application/json')}
function restore(file){const reader=new FileReader();reader.onload=()=>{try{const d=JSON.parse(reader.result);if(!Array.isArray(d.attempts))throw Error();state=Object.assign(defaults(),d);saveState();renderRop()}catch(e){alert('Файл резервной копии не распознан.')}};reader.readAsText(file)}
$('employeeMode').onclick=showStart;$('ropMode').onclick=renderRop;$('employeeId').oninput=checkEligibility;$('startButton').onclick=startExam;$('prevButton').onclick=()=>move(-1);$('nextButton').onclick=()=>move(1);$('saveSettings').onclick=saveSettings;$('exportCsv').onclick=csv;$('exportJson').onclick=backup;$('importJson').onchange=e=>e.target.files[0]&&restore(e.target.files[0]);showStart();
</script>
</body></html>'''


def build(bank_path: Path = BANK, output_path: Path = OUTPUT) -> Path:
    questions = load_bank(bank_path)
    data = json.dumps(questions, ensure_ascii=False, separators=(",", ":")).replace("</script", "<\\/script")
    output_path.write_text(HTML_TEMPLATE.replace("__QUESTION_DATA__", data), encoding="utf-8")
    return output_path


if __name__ == "__main__":
    path = build()
    print(path)
    print({"questions": len(load_bank(BANK)), "bytes": path.stat().st_size, "autonomous": True})
