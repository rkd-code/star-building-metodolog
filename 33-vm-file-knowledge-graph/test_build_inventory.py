import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

MODULE_PATH = Path(__file__).with_name("build_inventory.py")


def load_module():
    spec = importlib.util.spec_from_file_location("build_inventory", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class InventoryTests(unittest.TestCase):
    def test_classifies_numbered_iteration_and_portfolio_file(self):
        module = load_module()
        self.assertEqual("P-011", module.classify_project(Path("/home/roman/30-employee-testing-html/index.html"))["id"])
        self.assertEqual("PORTFOLIO", module.classify_project(Path("/home/roman/DECISIONS.md"))["id"])
        self.assertEqual("P-002", module.classify_project(Path("/home/roman/knowledge_base/full_REG-001.md"))["id"])

    def test_redacts_credentials_without_copying_values(self):
        module = load_module()
        secret = "abc123-DO-NOT-COPY"
        result, redactions = module.redact_text(f"api_key = '{secret}'\nОбычный текст")
        self.assertNotIn(secret, result)
        self.assertIn("[REDACTED]", result)
        self.assertGreater(redactions, 0)

    def test_excluded_directory_is_recorded_once_not_recursively(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            hidden = root / ".cache" / "nested"
            hidden.mkdir(parents=True)
            (hidden / "one.txt").write_text("one", encoding="utf-8")
            (hidden / "two.txt").write_text("two", encoding="utf-8")
            records, excluded = module.scan(root)
            self.assertEqual([], records)
            self.assertEqual(1, len(excluded))
            self.assertEqual(str(root / ".cache"), excluded[0]["path"])

    def test_build_writes_table_jsonl_and_integrity_safe_graph(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp) / "home"
            (root / "30-employee-testing-html").mkdir(parents=True)
            (root / "30-employee-testing-html" / "README.md").write_text("Платформа тестирования", encoding="utf-8")
            (root / "DECISIONS.md").write_text("Решения портфеля", encoding="utf-8")
            output = Path(temp) / "out"
            report = module.build_inventory(root=root, output_dir=output)
            self.assertEqual(2, report["indexed_files"])
            self.assertTrue((output / "file_inventory.xlsx").exists())
            self.assertTrue((output / "file_inventory.jsonl").exists())
            graph = json.loads((output / "file_graph.json").read_text(encoding="utf-8"))
            ids = {node["id"] for node in graph["nodes"]}
            self.assertIn("project:P-011", ids)
            self.assertIn("project:PORTFOLIO", ids)
            self.assertTrue(all(edge["source"] in ids and edge["target"] in ids for edge in graph["edges"]))
            workbook = load_workbook(output / "file_inventory.xlsx", read_only=True)
            self.assertEqual(["Файлы", "Проекты", "Исключения", "Сводка"], workbook.sheetnames)
            self.assertEqual(3, workbook["Файлы"].max_row)


if __name__ == "__main__":
    unittest.main()
