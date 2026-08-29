#!/usr/bin/env python3
"""Строит безопасный реестр рабочих файлов и граф их проектной принадлежности."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path("/home/roman")
OUTPUT_DIR = ROOT / "33-vm-file-knowledge-graph" / "output"
TEXT_EXTENSIONS = {".md", ".txt", ".py", ".html", ".htm", ".json", ".yaml", ".yml", ".toml", ".csv", ".sh", ".sql", ".js", ".ts", ".css", ".xml"}
DOCUMENT_EXTENSIONS = {".xlsx", ".docx"}
EXCLUDED_PARTS = {
    ".git", ".hermes", ".cache", ".local", ".npm", ".cargo", ".rustup",
    ".ssh", ".gnupg", ".config", ".codex", ".venv", "venv", "node_modules",
    "__pycache__", "Hermes-переписка", "graphify-out",
}
EXCLUDED_FILES = {".env", ".gitconfig", ".bash_history", "hmodel.log", "hermes-install.log"}
GENERATED_NAMES = {"file_inventory.xlsx", "file_inventory.csv", "file_inventory.jsonl", "file_graph.json", "file_graph.html", "inventory_report.json", "excluded_files.csv"}
MAX_XLSX_CONTENT = 32000
MAX_TEXT_BYTES = 5_000_000

PROJECTS = {
    "PORTFOLIO": "Общие исходные данные и управление портфелем Star Building",
    "P-001": "Корпоративная система «Кодификатор» Star Building",
    "P-002": "Нормативная база Star Building",
    "P-003": "«Смотритель» — кодификатор регламентов",
    "P-004": "«Ом» — корпоративный консультант",
    "P-005": "«Сверщик» и сквозной аудит",
    "P-006": "Интеграция Битрикс24",
    "P-007": "Контур приема материалов и Telegram",
    "P-008": "Маркетинговые агенты Star Building",
    "P-009": "Профили Hermes и управление разработкой",
    "P-010": "Единый граф проектов и знаний",
    "P-011": "«Кворум» — тестирование сотрудников",
    "P-012": "Архив переписки Hermes",
    "P-013": "Реестр файлов виртуальной машины",
}

PREFIX_PROJECT = {
    "01-report": "P-001", "02-tests": "P-001", "03-data-model": "P-001",
    "04-environment": "P-001", "05-integrations": "P-006", "06-tools": "P-001",
    "07-concept": "P-001", "08-refinement": "P-001", "09-roadmap": "P-001",
    "10-general-assistant": "P-004", "11-updated-artifacts": "P-001",
    "12-mindmap-update": "P-001", "14-intake-tool": "P-007",
    "15-bitrix24-integration": "P-006", "16-telegram-setup": "P-007",
    "17-two-bots-split": "P-003", "18-bot-behavior-refinement": "P-003",
    "19-top-20-roadmap": "P-002", "20-cross-check-engine": "P-005",
    "21-cross-audit": "P-005", "22-actual-master-package": "P-001",
    "23-code-names": "P-001", "24-marketing-bots": "P-008",
    "25-marketing-bots-detailed": "P-008", "26-hermes-profiles": "P-009",
    "27-developer-skills": "P-009", "28-unified-project-graph": "P-010",
    "29-employee-testing-service": "P-011", "30-employee-testing-html": "P-011",
    "31-employee-testing-html-v2": "P-011", "32-chat-history-archive": "P-012",
    "33-vm-file-knowledge-graph": "P-013", "knowledge_base": "P-002",
}

SECRET_PATTERNS = [
    re.compile(r"(?im)(api[_-]?key|token|password|secret|webhook|bot[_-]?token)(\s*[=:]\s*)(['\"]?)([^\s'\";,]+)(\3)"),
    re.compile(r"\b\d{8,12}:[A-Za-z0-9_-]{20,}\b"),
    re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----.*?-----END (?:RSA |EC |OPENSSH )?PRIVATE KEY-----", re.S),
]


def classify_project(path: Path) -> dict[str, str]:
    parts = path.parts
    if "knowledge_base" in parts:
        project_id = "P-002"
    else:
        project_id = next((PREFIX_PROJECT[part] for part in parts if part in PREFIX_PROJECT), "PORTFOLIO")
    return {"id": project_id, "name": PROJECTS[project_id]}


def redact_text(text: str) -> tuple[str, int]:
    total = 0
    result = text
    for index, pattern in enumerate(SECRET_PATTERNS):
        if index == 0:
            result, count = pattern.subn(lambda m: f"{m.group(1)}{m.group(2)}[REDACTED]", result)
        else:
            result, count = pattern.subn("[REDACTED]", result)
        total += count
    return result, total


def extract_xlsx(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    chunks = []
    for sheet in workbook.worksheets:
        chunks.append(f"### Лист: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                chunks.append("\t".join(values))
    workbook.close()
    return "\n".join(chunks)


def extract_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs = []
    for paragraph in root.iter(namespace + "p"):
        text = "".join(node.text or "" for node in paragraph.iter(namespace + "t"))
        if text.strip():
            paragraphs.append(text)
    return "\n".join(paragraphs)


def extract_content(path: Path) -> tuple[str, str]:
    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS or path.name == ".gitignore":
        if path.stat().st_size > MAX_TEXT_BYTES:
            return "", "исключен: текстовый файл больше 5 МБ"
        return path.read_text(encoding="utf-8", errors="replace"), "полный текст"
    if suffix == ".xlsx":
        return extract_xlsx(path), "извлечены значения листов"
    if suffix == ".docx":
        return extract_docx(path), "извлечен текст документа"
    return "", "исключен: неподдерживаемый или двоичный формат"


def should_exclude(path: Path, root: Path) -> str | None:
    relative = path.relative_to(root)
    if any(part in EXCLUDED_PARTS for part in relative.parts):
        return "служебный, секретный, кэшированный или зависимый каталог"
    if path.name in EXCLUDED_FILES:
        return "учетные, журнальные или системные данные"
    if path.name in GENERATED_NAMES:
        return "генерируемый результат инвентаризации"
    if path.is_symlink():
        return "символическая ссылка"
    if not path.is_file():
        return "не файл"
    if path.suffix.lower() not in TEXT_EXTENSIONS | DOCUMENT_EXTENSIONS and path.name != ".gitignore":
        return "неподдерживаемый или двоичный формат"
    return None


def describe_content(content: str, filename: str) -> str:
    plain = re.sub(r"[`#>*_\-]+", " ", content)
    plain = re.sub(r"\s+", " ", plain).strip()
    return (plain[:280] + ("…" if len(plain) > 280 else "")) or f"Файл {filename} без извлеченного текста"


def scan(root: Path) -> tuple[list[dict], list[dict]]:
    records, excluded = [], []
    candidate_files: list[Path] = []
    for directory, dirnames, filenames in os.walk(root, topdown=True, followlinks=False):
        base = Path(directory)
        kept_directories = []
        for dirname in sorted(dirnames):
            child = base / dirname
            if dirname in EXCLUDED_PARTS:
                excluded.append({"path": str(child), "name": dirname, "reason": "служебный, секретный, кэшированный или зависимый каталог"})
            else:
                kept_directories.append(dirname)
        dirnames[:] = kept_directories
        candidate_files.extend(base / filename for filename in sorted(filenames))

    for path in sorted(candidate_files):
        try:
            reason = should_exclude(path, root)
            if reason:
                excluded.append({"path": str(path), "name": path.name, "reason": reason})
                continue
            raw_content, extraction = extract_content(path)
            if extraction.startswith("исключен"):
                excluded.append({"path": str(path), "name": path.name, "reason": extraction})
                continue
            content, redactions = redact_text(raw_content)
            stat = path.stat()
            project = classify_project(path)
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            records.append({
                "path": str(path), "name": path.name, "extension": path.suffix.lower() or "без расширения",
                "project_id": project["id"], "project_name": project["name"],
                "content_type": extraction, "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "sha256": digest, "summary": describe_content(content, path.name),
                "content": content, "redactions": redactions,
            })
        except (OSError, PermissionError, ValueError, zipfile.BadZipFile) as error:
            excluded.append({"path": str(path), "name": path.name, "reason": f"ошибка обработки: {type(error).__name__}"})
    return records, excluded


def write_jsonl(records: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8") as stream:
        for record in records:
            stream.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_csv(records: list[dict], path: Path) -> None:
    fields = ["path", "name", "project_id", "project_name", "extension", "size_bytes", "modified_at", "sha256", "summary", "content", "redactions"]
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore", delimiter=";")
        writer.writeheader(); writer.writerows(records)


def style_sheet(sheet, widths: dict[str, int]) -> None:
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="173F35"); cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    for column, width in widths.items(): sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_xlsx(records: list[dict], excluded: list[dict], path: Path) -> None:
    workbook = Workbook(); files = workbook.active; files.title = "Файлы"
    headers = ["Путь", "Название", "Проект", "Название проекта", "Тип", "Размер, байт", "Изменен", "SHA-256", "Краткое содержание", "Содержимое", "Редакции секретов"]
    files.append(headers)
    for record in records:
        content = record["content"]
        if len(content) > MAX_XLSX_CONTENT: content = content[:MAX_XLSX_CONTENT] + "\n[СОДЕРЖИМОЕ СОКРАЩЕНО В EXCEL; ПОЛНЫЙ ТЕКСТ — В JSONL]"
        files.append([record["path"], record["name"], record["project_id"], record["project_name"], record["extension"], record["size_bytes"], record["modified_at"], record["sha256"], record["summary"], content, record["redactions"]])
    style_sheet(files, {"A":55,"B":28,"C":12,"D":42,"E":14,"F":14,"G":24,"H":18,"I":48,"J":90,"K":16})
    if files.max_row > 1:
        table = Table(displayName="FileInventory", ref=files.dimensions); table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True); files.add_table(table)

    projects = workbook.create_sheet("Проекты"); projects.append(["Код", "Проект", "Количество файлов", "Объем, байт"])
    counts, sizes = Counter(r["project_id"] for r in records), Counter()
    for record in records: sizes[record["project_id"]] += record["size_bytes"]
    for project_id in sorted(counts): projects.append([project_id, PROJECTS[project_id], counts[project_id], sizes[project_id]])
    style_sheet(projects, {"A":14,"B":55,"C":20,"D":20})

    omissions = workbook.create_sheet("Исключения"); omissions.append(["Путь", "Название", "Причина"])
    for item in excluded: omissions.append([item["path"], item["name"], item["reason"]])
    style_sheet(omissions, {"A":70,"B":35,"C":55})

    summary = workbook.create_sheet("Сводка"); summary.append(["Показатель", "Значение"])
    summary_rows = [
        ("Корневая папка", str(ROOT)), ("Проиндексировано файлов", len(records)),
        ("Исключено файлов", len(excluded)), ("Проектов", len(counts)),
        ("Общий объем индексированных файлов, байт", sum(r["size_bytes"] for r in records)),
        ("Скрыто фрагментов с признаками секретов", sum(r["redactions"] for r in records)),
        ("Примечание", "Полный очищенный текст хранится в file_inventory.jsonl; Excel ограничивает ячейку 32767 знаками."),
    ]
    for row in summary_rows: summary.append(row)
    style_sheet(summary, {"A":48,"B":95})
    workbook.save(path)


def build_graph(records: list[dict]) -> dict:
    project_ids = sorted({record["project_id"] for record in records})
    nodes = [{"id": f"project:{pid}", "label": f"{pid} — {PROJECTS[pid]}", "type": "project", "project_id": pid} for pid in project_ids]
    edges = []
    for record in records:
        node_id = "file:" + hashlib.sha256(record["path"].encode()).hexdigest()[:20]
        nodes.append({"id": node_id, "label": record["name"], "type": "file", "path": record["path"], "project_id": record["project_id"], "summary": record["summary"], "sha256": record["sha256"]})
        edges.append({"source": f"project:{record['project_id']}", "target": node_id, "relation": "CONTAINS_FILE"})
    return {"directed": True, "nodes": nodes, "edges": edges, "generated_at": datetime.now(timezone.utc).isoformat()}


def write_graph_html(graph: dict, path: Path) -> None:
    data = json.dumps(graph, ensure_ascii=False).replace("</script", "<\\/script")
    template = """<!doctype html><html lang='ru'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Граф файлов Star Building</title><style>body{margin:0;font:14px Arial;background:#f3f1ea;color:#142b26}header{padding:18px 24px;background:#123f34;color:white}main{display:grid;grid-template-columns:320px 1fr;min-height:calc(100vh - 74px)}aside{padding:18px;border-right:1px solid #ccd5d0;overflow:auto}input{width:100%;padding:11px;box-sizing:border-box}#projects button{width:100%;text-align:left;margin:5px 0;padding:10px;border:1px solid #ccd5d0;background:white;border-radius:8px;cursor:pointer}.cards{padding:20px;display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:12px;align-content:start}.card{background:white;border:1px solid #d7ddd9;border-radius:12px;padding:14px}.card b{display:block;margin-bottom:6px}.card small{color:#64736f}.card p{overflow-wrap:anywhere}.empty{padding:30px}@media(max-width:760px){main{grid-template-columns:1fr}aside{border-right:0;border-bottom:1px solid #ccd5d0}}</style></head><body><header><b>Единый граф файлов Star Building</b><div id='stats'></div></header><main><aside><input id='search' placeholder='Поиск по файлам и содержанию'><div id='projects'></div></aside><section id='cards' class='cards'></section></main><script id='data' type='application/json'>__DATA__</script><script>const G=JSON.parse(document.getElementById('data').textContent),files=G.nodes.filter(n=>n.type==='file'),projects=G.nodes.filter(n=>n.type==='project');let selected='';const esc=s=>String(s||'').replace(/[&<>\"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}[c]));function render(){const q=document.getElementById('search').value.toLowerCase(),shown=files.filter(f=>(!selected||f.project_id===selected)&&(!q||(f.label+' '+f.path+' '+f.summary).toLowerCase().includes(q)));document.getElementById('cards').innerHTML=shown.length?shown.map(f=>`<article class='card'><b>${esc(f.label)}</b><small>${esc(f.project_id)}</small><p>${esc(f.summary)}</p><small>${esc(f.path)}</small></article>`).join(''):`<div class='empty'>Файлы не найдены</div>`;document.getElementById('stats').textContent=`${files.length} файлов · ${projects.length} проектов · показано ${shown.length}`};document.getElementById('projects').innerHTML=`<button data-id=''>Все проекты</button>`+projects.map(p=>`<button data-id='${p.project_id}'>${esc(p.label)}</button>`).join('');document.querySelectorAll('#projects button').forEach(b=>b.onclick=()=>{selected=b.dataset.id;render()});document.getElementById('search').oninput=render;render();</script></body></html>"""
    path.write_text(template.replace("__DATA__", data), encoding="utf-8")


def build_inventory(root: Path = ROOT, output_dir: Path = OUTPUT_DIR) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    records, excluded = scan(root)
    write_jsonl(records, output_dir / "file_inventory.jsonl")
    write_csv(records, output_dir / "file_inventory.csv")
    with (output_dir / "excluded_files.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["path", "name", "reason"], delimiter=";"); writer.writeheader(); writer.writerows(excluded)
    write_xlsx(records, excluded, output_dir / "file_inventory.xlsx")
    graph = build_graph(records)
    (output_dir / "file_graph.json").write_text(json.dumps(graph, ensure_ascii=False, indent=2), encoding="utf-8")
    write_graph_html(graph, output_dir / "file_graph.html")
    report = {
        "root": str(root), "indexed_files": len(records), "excluded_files": len(excluded),
        "projects": len({r['project_id'] for r in records}), "bytes": sum(r["size_bytes"] for r in records),
        "redactions": sum(r["redactions"] for r in records), "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    (output_dir / "inventory_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=ROOT); parser.add_argument("--output", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args(); print(json.dumps(build_inventory(args.root, args.output), ensure_ascii=False)); return 0


if __name__ == "__main__":
    raise SystemExit(main())
