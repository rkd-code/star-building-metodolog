#!/usr/bin/env python3
"""Создает проверяемый Excel-шаблон банка вопросов."""

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).with_name("question_bank_template.xlsx")

HEADERS = {
    "Вопросы": ["Код вопроса", "Тема", "Тип", "Сложность", "Формулировка", "Объяснение", "Код источника", "Пункт источника", "Статус", "Версия", "Автор", "Проверяющий", "Дата проверки"],
    "Варианты": ["Код вопроса", "Код варианта", "Текст варианта", "Правильный", "Вес", "Порядок"],
    "Назначение_по_должности": ["Код вопроса", "Официальная должность", "Обязательность", "Комментарий"],
    "Программы": ["Код программы", "Название", "Официальная должность", "Темы", "Число вопросов", "Время, минут", "Порог, %", "Попыток", "Срок результата, месяцев", "Статус", "Версия"],
    "Источники": ["Код источника", "Название", "Версия", "Статус", "Путь или ссылка", "Дата проверки"],
    "Справочники": ["Типы вопросов", "Сложность", "Статусы вопроса", "Да/Нет", "Обязательность"],
    "Инструкция": ["Раздел", "Правило"],
}

ROWS = {
    "Справочники": [
        ["Один ответ", "Базовый", "Черновик", "Да", "Обязательный"],
        ["Несколько ответов", "Средний", "На проверке", "Нет", "Дополнительный"],
        ["Верно/неверно", "Повышенный", "Действует", None, None],
        ["Последовательность", None, "Приостановлен", None, None],
        ["Ситуационная задача", None, "Архив", None, None],
    ],
    "Инструкция": [
        ["Назначение", "Шаблон рассчитан минимум на 500 вопросов; пустые строки не являются вопросами."],
        ["Достоверность", "Действующий вопрос обязан ссылаться на утвержденный документ и конкретный пункт."],
        ["Безопасность", "Не сохраняйте пароли, токены, ключи и персональные секреты."],
        ["Версионность", "Изменение правильного ответа создает новую версию вопроса."],
        ["Должности", "Используйте только официальные должности из утвержденной оргструктуры."],
        ["Числа", "Порог, время, попытки и срок результата утверждаются владельцем процесса."],
    ],
}


def style_sheet(ws, widths):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width


def add_table(ws, name):
    ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


def build():
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers in HEADERS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in ROWS.get(name, []):
            ws.append(row)
        if name not in ROWS:
            # Одна пустая строка нужна для корректной таблицы; пользователь расширяет ее при заполнении.
            ws.append([None] * len(headers))
        style_sheet(ws, [20] * len(headers))
        add_table(ws, "Таблица" + str(len(wb.worksheets)))

    questions = wb["Вопросы"]
    dv_type = DataValidation(type="list", formula1="='Справочники'!$A$2:$A$6")
    dv_level = DataValidation(type="list", formula1="='Справочники'!$B$2:$B$4")
    dv_status = DataValidation(type="list", formula1="='Справочники'!$C$2:$C$6")
    questions.add_data_validation(dv_type); dv_type.add("C2:C501")
    questions.add_data_validation(dv_level); dv_level.add("D2:D501")
    questions.add_data_validation(dv_status); dv_status.add("I2:I501")

    variants = wb["Варианты"]
    dv_yesno = DataValidation(type="list", formula1="='Справочники'!$D$2:$D$3")
    variants.add_data_validation(dv_yesno); dv_yesno.add("D2:D3001")

    mapping = wb["Назначение_по_должности"]
    dv_required = DataValidation(type="list", formula1="='Справочники'!$E$2:$E$3")
    mapping.add_data_validation(dv_required); dv_required.add("C2:C2001")

    wb.calculation.fullCalcOnLoad = True
    wb.save(OUT)
    return OUT


def verify(path):
    wb = load_workbook(path, read_only=False, data_only=False)
    assert wb.sheetnames == list(HEADERS)
    assert wb["Вопросы"].max_column == len(HEADERS["Вопросы"])
    assert len(wb["Вопросы"].data_validations.dataValidation) == 3
    assert len(wb["Варианты"].data_validations.dataValidation) == 1
    assert len(wb["Назначение_по_должности"].data_validations.dataValidation) == 1
    return {"листы": wb.sheetnames, "емкость_вопросов": 500, "емкость_вариантов": 3000}


if __name__ == "__main__":
    path = build()
    print(path)
    print(verify(path))
